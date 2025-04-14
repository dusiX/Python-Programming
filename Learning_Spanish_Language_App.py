import openpyxl
from random import randint
from PyQt6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QLineEdit, QDialog, QDialogButtonBox
)
import sys
import os

user_folder = os.path.expanduser('~')
file_path = os.path.join(user_folder, 'Desktop', 'Language_App', 'Words.xlsx')

wb = openpyxl.load_workbook(file_path)
sheet = wb['Wszystkie']

class InputDialog(QDialog):
    def __init__(self, title, body="Co chcesz zrobić? :)", parent=None):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setFixedSize(250, 100)

        self.body_label = QLabel(body)
        self.body_label.setWordWrap(True)

        self.input_field = QLineEdit(self)
        self.input_field.setPlaceholderText("Podaj tłumaczenie")

        if title == "Przetłumacz słowo z polskiego na hiszpański":
            self.buttons_layout = QHBoxLayout()
            self.add_button("á")
            self.add_button("é")
            self.add_button("í")
            self.add_button("ó")
            self.add_button("ú")
            self.add_button("ñ")

        self.buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok,
            self
        )
        self.buttons.accepted.connect(self.accept)
        self.buttons.rejected.connect(self.reject)

        layout = QVBoxLayout()
        layout.addWidget(self.body_label)
        layout.addWidget(self.input_field)
        if title == "Przetłumacz słowo z polskiego na hiszpański":
            layout.addLayout(self.buttons_layout)
        layout.addWidget(self.buttons)
        self.setLayout(layout)
    
    def add_button(self, char):
        button = QPushButton(char, self)
        button.setFixedSize(30, 30)
        button.clicked.connect(lambda: self.on_button_clicked(char))
        self.buttons_layout.addWidget(button)

    def on_button_clicked(self, char):
        current_text = self.input_field.text()
        self.input_field.setText(current_text + char)

    def get_input(self):
        return self.input_field.text()

class ResultDialog(QDialog):
    def __init__(self, result_message, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Wynik")
        self.setFixedSize(300, 150)

        self.result_label = QLabel(result_message, self)
        self.result_label.setWordWrap(True)

        self.buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok, self)
        self.buttons.accepted.connect(self.accept)

        layout = QVBoxLayout()
        layout.addWidget(self.result_label)
        layout.addWidget(self.buttons)
        self.setLayout(layout)

class WordEntryDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Wpisz nowe słowa")
        self.setFixedSize(300, 200)

        self.spanish_label = QLabel("Wpisz słowo po hiszpańsku:")
        self.polish_label = QLabel("Wpisz tłumaczenie po polsku:")

        self.spanish_input = QLineEdit(self)
        self.spanish_input.setPlaceholderText("Słowo po hiszpańsku")

        self.polish_input = QLineEdit(self)
        self.polish_input.setPlaceholderText("Tłumaczenie po polsku")

        self.buttons_layout = QHBoxLayout()
        self.add_button("á")
        self.add_button("é")
        self.add_button("í")
        self.add_button("ó")
        self.add_button("ú")
        self.add_button("ñ")

        self.buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel, self)
        self.buttons.accepted.connect(self.accept)
        self.buttons.rejected.connect(self.reject)

        layout = QVBoxLayout()
        layout.addWidget(self.spanish_label)
        layout.addWidget(self.spanish_input)
        layout.addLayout(self.buttons_layout)
        layout.addWidget(self.polish_label)
        layout.addWidget(self.polish_input)
        layout.addWidget(self.buttons)

        self.setLayout(layout)

    def add_button(self, char):
        button = QPushButton(char, self)
        button.setFixedSize(30, 30)
        button.clicked.connect(lambda: self.on_button_clicked(char))
        self.buttons_layout.addWidget(button)

    def on_button_clicked(self, char):
        current_text = self.spanish_input.text()
        self.spanish_input.setText(current_text + char)

    def get_inputs(self):
        return self.spanish_input.text(), self.polish_input.text()

class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Fiszki")
        self.setFixedSize(300, 150)

        layout = QVBoxLayout(self)

        btn1 = QPushButton("Tłumacz z polskiego na hiszpański")
        btn1.clicked.connect(self.action1)
        layout.addWidget(btn1)

        btn2 = QPushButton("Tłumacz z hiszpańskiego na polski")
        btn2.clicked.connect(self.action2)
        layout.addWidget(btn2)

        btn3 = QPushButton("Dodaj nowe słowo")
        btn3.clicked.connect(self.action3)
        layout.addWidget(btn3)

        self.setLayout(layout)

    def action1(self):
        learn_words = []
        repeat_words = []
        learned_words = []

        for row in range(2, sheet.max_row + 1):
            status = sheet.cell(row, 4).value
            if status == "DO NAUCZENIA":
                learn_words.append(row)
            elif status == "DO POWTÓRKI":
                repeat_words.append(row)
            elif status == "NAUCZONE":
                learned_words.append(row)

        category = randint(1, 100)
        if category <= 60 and learn_words:
            chosen_row = learn_words[randint(0, len(learn_words) - 1)]
        elif category <= 90 and repeat_words:
            chosen_row = repeat_words[randint(0, len(repeat_words) - 1)]
        elif learned_words:
            chosen_row = learned_words[randint(0, len(learned_words) - 1)]
        elif repeat_words:
            chosen_row = repeat_words[randint(0, len(repeat_words) - 1)]
        elif learn_words:
            chosen_row = learn_words[randint(0, len(learn_words) - 1)]
        else:
            chosen_row = randint(2, sheet.max_row)

        word_to_translate = get_value(2, chosen_row)
        translated_word = get_value(1, chosen_row)
        dialog = InputDialog("Przetłumacz słowo z polskiego na hiszpański","Słowo do przetłumaczenia: \n\n" + word_to_translate + "\n\nPodaj przetłumaczone słowo:", self)
        if dialog.exec():
            if translated_word == dialog.get_input():
                result_message = "Poprawna odpowiedź!"
            else:
                result_message = "Błędna odpowiedź! Prawidłowa odpowiedź to:\n\n" + translated_word
            
            result_dialog = ResultDialog(result_message + "\n\nCzy chcesz spróbować następnego słowa?", self)

            result_dialog.buttons.setStandardButtons(QDialogButtonBox.StandardButton.Ok)

            set_value('D', chosen_row, translated_word == dialog.get_input())

            if result_dialog.exec() == QDialog.DialogCode.Accepted:
                self.action1()


    def action2(self):
        learn_words = []
        repeat_words = []
        learned_words = []

        for row in range(1, sheet.max_row + 1):
            status = sheet.cell(row, 3).value
            if status == "DO NAUCZENIA":
                learn_words.append(row)
            elif status == "DO POWTÓRKI":
                repeat_words.append(row)
            elif status == "NAUCZONE":
                learned_words.append(row)

        category = randint(1, 100)
        if category <= 60 and learn_words:
            chosen_row = learn_words[randint(0, len(learn_words) - 1)]
        elif category <= 90 and repeat_words:
            chosen_row = repeat_words[randint(0, len(repeat_words) - 1)]
        elif learned_words:
            chosen_row = learned_words[randint(0, len(learned_words) - 1)]
        elif repeat_words:
            chosen_row = repeat_words[randint(0, len(repeat_words) - 1)]
        elif learn_words:
            chosen_row = learn_words[randint(0, len(learn_words) - 1)]
        else:
            chosen_row = randint(1, sheet.max_row)

        word_to_translate = get_value(1, chosen_row)
        translated_word = get_value(2, chosen_row)
        dialog = InputDialog("Przetłumacz słowo z hiszpańskiego na polski","Słowo do przetłumaczenia: \n\n" + word_to_translate + "\n\nPodaj przetłumaczone słowo:", self)
        if dialog.exec():
            if translated_word == dialog.get_input():
                result_message = "Poprawna odpowiedź!"
            else:
                result_message = "Błędna odpowiedź! Prawidłowa odpowiedź to:\n\n" + translated_word
            
            result_dialog = ResultDialog(result_message + "\n\nCzy chcesz spróbować następnego słowa?", self)

            result_dialog.buttons.setStandardButtons(QDialogButtonBox.StandardButton.Ok)

            set_value('C', chosen_row, translated_word == dialog.get_input())

            if result_dialog.exec() == QDialog.DialogCode.Accepted:
                self.action2()

    def action3(self):
        row = sheet.max_row + 1
        new_polish_word_adress = 'B' + str(row)
        new_spanish_word_adress = 'A' + str(row)

        dialog = WordEntryDialog(self)
        if dialog.exec():
            spanish_word, polish_translation = dialog.get_inputs()

        if dialog.get_inputs():
            sheet[new_polish_word_adress] = polish_translation
            sheet[new_spanish_word_adress] = spanish_word
            sheet['C' + str(row)] = 'DO NAUCZENIA'
            sheet['D' + str(row)] = 'DO NAUCZENIA'

        wb.save(file_path)

def get_value(column, row):
    return sheet.cell(row, column).value

def set_value(column, row, correct):
    adress = column + str(row)
    if correct == True:
        match(sheet[adress].value):
            case 'NAUCZONE':
                pass
            case 'DO NAUCZENIA':
                sheet[adress] = 'DO POWTÓRKI'
            case 'DO POWTÓRKI':
                sheet[adress] = 'NAUCZONE'
    else:
        match(sheet[adress].value):
            case 'NAUCZONE':
                sheet[adress] = 'DO POWTÓRKI'
            case 'DO NAUCZENIA':
                pass
            case 'DO POWTÓRKI':
                sheet[adress] = 'DO NAUCZENIA'
    
    wb.save(file_path)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
